# -*- coding: utf-8 -*-
import pathlib

try:
    import openpyxl
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"])
    import openpyxl

base = pathlib.Path(r"D:\Git\Public_Trade_Module\Import date")

for f in sorted(base.glob("*.xlsx")):
    print(f"\n=== {f.name} ===")
    wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(max_row=5, values_only=True))
    for i, row in enumerate(rows):
        tag = "HEADER" if i == 0 else f"row {i+1}"
        print(f"  [{tag}] {list(row)}")
    print(f"  Total rows: {ws.max_row}")
    wb.close()
