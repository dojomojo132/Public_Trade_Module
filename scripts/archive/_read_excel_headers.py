# -*- coding: utf-8 -*-
"""Одноразовый скрипт: анализ данных Excel GreenMarket"""
import openpyxl
from pathlib import Path

fpath = Path(r"D:\Git\Public_Trade_Module\Выгрузка номенклатура GreenMarket.xlsx")
wb = openpyxl.load_workbook(fpath, read_only=True)
ws = wb.active

# Find rows with price and quantity
print("=== ROWS WITH PRICES/QTY ===")
count = 0
for row in ws.iter_rows(min_row=2, max_row=500):
    vals = [str(c.value) if c.value else "" for c in row]
    if vals[3] or vals[4] or vals[8]:  # Col 4,5,9
        print(f"Art={vals[0]}, Name={vals[1][:40]}, UoM={vals[2]}, Price={vals[3]}, Qty={vals[4]}, BC={vals[5]}, Tax={vals[6]}, UKTZED={vals[7]}, PurchP={vals[8]}")
        count += 1
        if count >= 15:
            break

# Total stats
total = has_name = has_bc = has_price = has_qty = has_pp = has_uom = 0
for row in ws.iter_rows(min_row=2, max_row=5000):
    vals = [str(c.value) if c.value else "" for c in row]
    if not vals[1]:
        continue
    total += 1
    if vals[1]: has_name += 1
    if vals[5]: has_bc += 1
    if vals[3]: has_price += 1
    if vals[4]: has_qty += 1
    if vals[8]: has_pp += 1
    if vals[2]: has_uom += 1

print(f"\n=== STATS ===")
print(f"total={total}, name={has_name}, bc={has_bc}, retail_price={has_price}, qty={has_qty}, purchase_price={has_pp}, uom={has_uom}")

wb.close()
