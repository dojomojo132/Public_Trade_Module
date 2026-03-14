# -*- coding: utf-8 -*-
"""Read Excel file to analyze structure for price import."""
import pathlib

try:
    import openpyxl
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--quiet"])
    import openpyxl

root = pathlib.Path(r"D:\Git\Public_Trade_Module")
xlsx = next(root.glob("*.xlsx"))
print(f"File: {xlsx.name}")

wb = openpyxl.load_workbook(xlsx, data_only=True)
ws = wb.active
print(f"Sheet: {ws.title}, Rows: {ws.max_row}, Cols: {ws.max_column}")
print()

for row in ws.iter_rows(min_row=1, max_row=20, values_only=False):
    vals = [(c.column_letter, c.value) for c in row if c.value is not None]
    if vals:
        print(vals)
wb.close()
