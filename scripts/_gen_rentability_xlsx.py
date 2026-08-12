# -*- coding: utf-8 -*-
"""Generate sample profitability Excel reports for PTM review."""
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.chart.label import DataLabelList
from collections import defaultdict
import random
from datetime import date, timedelta
from pathlib import Path

random.seed(42)

out = Path(r"D:\Git\Public_Trade_Module\Документация\Тестирование\Примеры_Отчёт_Рентабельность.xlsx")
out.parent.mkdir(parents=True, exist_ok=True)

wb = Workbook()

thin = Border(
    left=Side(style="thin", color="B0B8C4"),
    right=Side(style="thin", color="B0B8C4"),
    top=Side(style="thin", color="B0B8C4"),
    bottom=Side(style="thin", color="B0B8C4"),
)
header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF", size=11)
title_font = Font(bold=True, size=14, color="1F4E79")
subtitle_font = Font(italic=True, size=10, color="5A6A7A")
kpi_label = Font(size=9, color="5A6A7A")
kpi_value = Font(bold=True, size=16, color="1F4E79")
money_fmt = "#,##0.00"
pct_fmt = "0.0%"
num_fmt = "#,##0.000"
alt_row = PatternFill("solid", fgColor="F5F8FC")
total_fill = PatternFill("solid", fgColor="D6E3F0")
total_font = Font(bold=True, size=11)

products = [
    ("Молоко 2.5% 1л", "Молочные"),
    ("Хлеб белый 500г", "Хлеб"),
    ("Яйца С1 10шт", "Яйца"),
    ("Сыр Гауда 200г", "Молочные"),
    ("Колбаса варёная 400г", "Мясо"),
    ("Курица охлажд. кг", "Мясо"),
    ("Яблоки кг", "Овощи-фрукты"),
    ("Бананы кг", "Овощи-фрукты"),
    ("Картофель кг", "Овощи-фрукты"),
    ("Масло подсолн. 1л", "Бакалея"),
    ("Сахар 1кг", "Бакалея"),
    ("Рис 1кг", "Бакалея"),
    ("Кофе молотый 250г", "Напитки"),
    ("Чай чёрный 100пак", "Напитки"),
    ("Вода 1.5л", "Напитки"),
    ("Пиво 0.5л", "Алкоголь"),
    ("Сигареты премиум", "Табак"),
    ("Шоколад 90г", "Кондитерка"),
    ("Печенье 300г", "Кондитерка"),
    ("Йогурт 150г", "Молочные"),
]
cashiers = ["Касса №1 (зал)", "Касса №2 (зал)", "Касса №3 (табак)"]
staff = ["Иванова А.П.", "Петренко М.С.", "Ковальчук О.В.", "Сидоренко Т.И."]
base = date(2026, 6, 1)


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        cell.border = thin


def style_data_row(ws, row, cols, is_total=False):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = thin
        cell.alignment = Alignment(horizontal="center" if c > 2 else "left", vertical="center")
        if is_total:
            cell.fill = total_fill
            cell.font = total_font
        elif row % 2 == 0:
            cell.fill = alt_row


def auto_width(ws, min_w=10, max_w=28):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = 0
        for cell in col:
            if cell.value is not None:
                length = max(length, min(len(str(cell.value)), max_w))
        ws.column_dimensions[letter].width = max(min_w, length + 2)


def gen_rows(n=90):
    rows = []
    for _ in range(n):
        name, group = random.choice(products)
        d = base + timedelta(days=random.randint(0, 29))
        qty = round(random.uniform(0.5, 40), 3)
        cost_u = round(random.uniform(8, 220), 2)
        markup = random.uniform(1.08, 1.55)
        sell_u = round(cost_u * markup, 2)
        cost = round(cost_u * qty, 2)
        sell = round(sell_u * qty, 2)
        gp = round(sell - cost, 2)
        r = (gp / sell) if sell else 0
        rows.append(
            {
                "date": d,
                "month": d.strftime("%Y-%m"),
                "product": name,
                "group": group,
                "cashier": random.choice(cashiers),
                "staff": random.choice(staff),
                "qty": qty,
                "cost": cost,
                "sell": sell,
                "gp": gp,
                "r": r,
            }
        )
    return rows


rows = gen_rows(90)

# ---------- 00 description ----------
ws = wb.active
ws.title = "00_Описание"
ws["A1"] = "PTM — примеры отчётов о рентабельности (тестовые данные)"
ws["A1"].font = title_font
ws["A2"] = (
    "Случайные данные, seed=42. Не из живой ИБ. "
    "Формулы: Валовая прибыль = Продажи − Себестоимость; Рентабельность % = Валовая / Продажи."
)
ws["A2"].font = subtitle_font
ws.merge_cells("A2:F2")

desc = [
    ("Лист", "Назначение", "Группировки", "Метрики"),
    (
        "01_По_номенклатуре",
        "Зеркало текущего СКД «Валовая прибыль»",
        "День → Номенклатура",
        "Кол-во, Себест., Продажи, Вал.прибыль, Рент.%",
    ),
    (
        "02_По_кассам_группам",
        "Как «Анализ продаж», но с маржой",
        "Касса → Группа → Номенклатура",
        "те же 5 метрик + итоги",
    ),
    (
        "03_По_сотрудникам",
        "Мотивация / контроль смен",
        "Сотрудник → Номенклатура",
        "Продажи, Себест., Маржа, %",
    ),
    (
        "04_KPI_свод",
        "P&L-lite: только gross",
        "KPI + матрица по месяцам × группам",
        "Σ и %",
    ),
    (
        "05_Сырые_данные",
        "Исходник для pivot (можно крутить в Excel)",
        "плоская таблица",
        "все поля",
    ),
]
for i, row in enumerate(desc, start=4):
    for j, v in enumerate(row, start=1):
        cell = ws.cell(row=i, column=j, value=v)
        if i == 4:
            cell.fill = header_fill
            cell.font = header_font
        cell.border = thin
ws["A11"] = "Цветовая шкала на %: красный (низкая) → жёлтый → зелёный (высокая)."
ws["A12"] = "Рекомендация: смотреть листы 01–04 как прототипы макетов СКД; 05 — сводная таблица."
ws["A11"].font = subtitle_font
ws["A12"].font = subtitle_font
auto_width(ws)
ws.column_dimensions["B"].width = 45
ws.column_dimensions["C"].width = 35
ws.column_dimensions["D"].width = 40

# ---------- 01 by product ----------
ws1 = wb.create_sheet("01_По_номенклатуре")
ws1["A1"] = "Валовая прибыль по номенклатуре / периодам"
ws1["A1"].font = title_font
ws1["A2"] = "Аналог отчёта Анл_ВаловаяПрибыль. Фильтр: июнь 2026. Сортировка по валовой прибыли (убыв.)."
ws1["A2"].font = subtitle_font
ws1.merge_cells("A2:H2")

headers1 = [
    "Дата",
    "Номенклатура",
    "Группа",
    "Количество",
    "Себестоимость",
    "Продажная сумма",
    "Валовая прибыль",
    "Рентабельность %",
]
for j, h in enumerate(headers1, 1):
    ws1.cell(row=4, column=j, value=h)
style_header(ws1, 4, 8)
ws1.row_dimensions[4].height = 30

agg1 = defaultdict(lambda: {"qty": 0, "cost": 0, "sell": 0, "gp": 0, "group": ""})
for r in rows:
    k = (r["date"], r["product"])
    a = agg1[k]
    a["qty"] += r["qty"]
    a["cost"] += r["cost"]
    a["sell"] += r["sell"]
    a["gp"] += r["gp"]
    a["group"] = r["group"]

items1 = sorted(agg1.items(), key=lambda x: -x[1]["gp"])
row_i = 5
for (d, prod), a in items1:
    rent = a["gp"] / a["sell"] if a["sell"] else 0
    vals = [
        d,
        prod,
        a["group"],
        round(a["qty"], 3),
        round(a["cost"], 2),
        round(a["sell"], 2),
        round(a["gp"], 2),
        rent,
    ]
    for j, v in enumerate(vals, 1):
        cell = ws1.cell(row=row_i, column=j, value=v)
        if j == 1:
            cell.number_format = "DD.MM.YYYY"
        if j in (5, 6, 7):
            cell.number_format = money_fmt
        if j == 4:
            cell.number_format = num_fmt
        if j == 8:
            cell.number_format = pct_fmt
    style_data_row(ws1, row_i, 8)
    row_i += 1

tot_q = sum(a["qty"] for a in agg1.values())
tot_c = sum(a["cost"] for a in agg1.values())
tot_s = sum(a["sell"] for a in agg1.values())
tot_g = sum(a["gp"] for a in agg1.values())
for j, v in enumerate(
    ["", "ИТОГО", "", round(tot_q, 3), round(tot_c, 2), round(tot_s, 2), round(tot_g, 2), (tot_g / tot_s if tot_s else 0)],
    1,
):
    cell = ws1.cell(row=row_i, column=j, value=v)
    if j in (5, 6, 7):
        cell.number_format = money_fmt
    if j == 4:
        cell.number_format = num_fmt
    if j == 8:
        cell.number_format = pct_fmt
style_data_row(ws1, row_i, 8, is_total=True)

ws1.conditional_formatting.add(
    f"H5:H{row_i - 1}",
    ColorScaleRule(
        start_type="min",
        start_color="F8696B",
        mid_type="percentile",
        mid_value=50,
        mid_color="FFEB84",
        end_type="max",
        end_color="63BE7B",
    ),
)
auto_width(ws1)
ws1.auto_filter.ref = f"A4:H{row_i - 1}"
ws1.freeze_panes = "A5"

# ---------- 02 cashiers / groups ----------
ws2 = wb.create_sheet("02_По_кассам_группам")
ws2["A1"] = "Валовая прибыль по кассам / группам номенклатуры"
ws2["A1"].font = title_font
ws2["A2"] = "Структура как «Анализ продаж», плюс себестоимость и маржа. Иерархия: Касса → Группа → Номенклатура."
ws2["A2"].font = subtitle_font
ws2.merge_cells("A2:H2")

headers2 = [
    "Касса",
    "Группа",
    "Номенклатура",
    "Количество",
    "Себестоимость",
    "Продажная сумма",
    "Валовая прибыль",
    "Рентабельность %",
]
for j, h in enumerate(headers2, 1):
    ws2.cell(row=4, column=j, value=h)
style_header(ws2, 4, 8)

agg2 = defaultdict(lambda: {"qty": 0, "cost": 0, "sell": 0, "gp": 0})
for r in rows:
    k = (r["cashier"], r["group"], r["product"])
    a = agg2[k]
    a["qty"] += r["qty"]
    a["cost"] += r["cost"]
    a["sell"] += r["sell"]
    a["gp"] += r["gp"]

items2 = sorted(agg2.items(), key=lambda x: (x[0][0], x[0][1], -x[1]["gp"]))
row_i = 5
cur_cashier = None
cashier_sub = {"qty": 0, "cost": 0, "sell": 0, "gp": 0}
cashier_fill = PatternFill("solid", fgColor="BDD7EE")


def write_subtotal(ws, row, label, a, level_fill):
    rent = a["gp"] / a["sell"] if a["sell"] else 0
    vals = [label, "", "", round(a["qty"], 3), round(a["cost"], 2), round(a["sell"], 2), round(a["gp"], 2), rent]
    for j, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=j, value=v)
        cell.fill = level_fill
        cell.font = total_font
        cell.border = thin
        if j in (5, 6, 7):
            cell.number_format = money_fmt
        if j == 4:
            cell.number_format = num_fmt
        if j == 8:
            cell.number_format = pct_fmt
    return row + 1


for (cash, grp, prod), a in items2:
    if cur_cashier is not None and cash != cur_cashier:
        row_i = write_subtotal(ws2, row_i, f"Итого: {cur_cashier}", cashier_sub, cashier_fill)
        cashier_sub = {"qty": 0, "cost": 0, "sell": 0, "gp": 0}
    cur_cashier = cash
    for k in cashier_sub:
        cashier_sub[k] += a[k]
    rent = a["gp"] / a["sell"] if a["sell"] else 0
    vals = [cash, grp, prod, round(a["qty"], 3), round(a["cost"], 2), round(a["sell"], 2), round(a["gp"], 2), rent]
    for j, v in enumerate(vals, 1):
        cell = ws2.cell(row=row_i, column=j, value=v)
        if j in (5, 6, 7):
            cell.number_format = money_fmt
        if j == 4:
            cell.number_format = num_fmt
        if j == 8:
            cell.number_format = pct_fmt
    style_data_row(ws2, row_i, 8)
    row_i += 1
if cur_cashier:
    row_i = write_subtotal(ws2, row_i, f"Итого: {cur_cashier}", cashier_sub, cashier_fill)

gt = {
    "qty": sum(a["qty"] for a in agg2.values()),
    "cost": sum(a["cost"] for a in agg2.values()),
    "sell": sum(a["sell"] for a in agg2.values()),
    "gp": sum(a["gp"] for a in agg2.values()),
}
row_i = write_subtotal(ws2, row_i, "ВСЕГО", gt, total_fill)

ws2.conditional_formatting.add(
    f"H5:H{row_i - 1}",
    ColorScaleRule(
        start_type="min",
        start_color="F8696B",
        mid_type="percentile",
        mid_value=50,
        mid_color="FFEB84",
        end_type="max",
        end_color="63BE7B",
    ),
)
auto_width(ws2)
ws2.freeze_panes = "A5"

# ---------- 03 staff ----------
ws3 = wb.create_sheet("03_По_сотрудникам")
ws3["A1"] = "Валовая прибыль по сотрудникам (кассирам)"
ws3["A1"].font = title_font
ws3["A2"] = "Срез по измерению Сотрудник регистра Продажи. Полезно для мотивации и контроля смен."
ws3["A2"].font = subtitle_font

for j, h in enumerate(
    ["Сотрудник", "Количество", "Себестоимость", "Продажная сумма", "Валовая прибыль", "Рентабельность %", "Доля продаж %"],
    1,
):
    ws3.cell(row=4, column=j, value=h)
style_header(ws3, 4, 7)

agg_staff = defaultdict(lambda: {"qty": 0, "cost": 0, "sell": 0, "gp": 0})
for r in rows:
    a = agg_staff[r["staff"]]
    a["qty"] += r["qty"]
    a["cost"] += r["cost"]
    a["sell"] += r["sell"]
    a["gp"] += r["gp"]

tot_sell_all = sum(a["sell"] for a in agg_staff.values())
row_i = 5
for st, a in sorted(agg_staff.items(), key=lambda x: -x[1]["gp"]):
    rent = a["gp"] / a["sell"] if a["sell"] else 0
    share = a["sell"] / tot_sell_all if tot_sell_all else 0
    vals = [st, round(a["qty"], 3), round(a["cost"], 2), round(a["sell"], 2), round(a["gp"], 2), rent, share]
    for j, v in enumerate(vals, 1):
        cell = ws3.cell(row=row_i, column=j, value=v)
        if j in (3, 4, 5):
            cell.number_format = money_fmt
        if j == 2:
            cell.number_format = num_fmt
        if j in (6, 7):
            cell.number_format = pct_fmt
    style_data_row(ws3, row_i, 7)
    row_i += 1

chart = BarChart()
chart.type = "col"
chart.title = "Валовая прибыль по сотрудникам"
chart.y_axis.title = "грн"
data = Reference(ws3, min_col=5, min_row=4, max_row=row_i - 1)
cats = Reference(ws3, min_col=1, min_row=5, max_row=row_i - 1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.shape = 4
chart.width = 15
chart.height = 8
ws3.add_chart(chart, "I4")

row_i += 2
ws3.cell(row=row_i, column=1, value="Детализация: сотрудник → номенклатура").font = Font(
    bold=True, size=12, color="1F4E79"
)
row_i += 1
for j, h in enumerate(
    ["Сотрудник", "Номенклатура", "Количество", "Себестоимость", "Продажная сумма", "Валовая прибыль", "Рентабельность %"],
    1,
):
    ws3.cell(row=row_i, column=j, value=h)
style_header(ws3, row_i, 7)
row_i += 1

agg3 = defaultdict(lambda: {"qty": 0, "cost": 0, "sell": 0, "gp": 0})
for r in rows:
    k = (r["staff"], r["product"])
    a = agg3[k]
    a["qty"] += r["qty"]
    a["cost"] += r["cost"]
    a["sell"] += r["sell"]
    a["gp"] += r["gp"]
for (st, prod), a in sorted(agg3.items(), key=lambda x: (x[0][0], -x[1]["gp"])):
    rent = a["gp"] / a["sell"] if a["sell"] else 0
    vals = [st, prod, round(a["qty"], 3), round(a["cost"], 2), round(a["sell"], 2), round(a["gp"], 2), rent]
    for j, v in enumerate(vals, 1):
        cell = ws3.cell(row=row_i, column=j, value=v)
        if j in (4, 5, 6):
            cell.number_format = money_fmt
        if j == 3:
            cell.number_format = num_fmt
        if j == 7:
            cell.number_format = pct_fmt
    style_data_row(ws3, row_i, 7)
    row_i += 1

auto_width(ws3)
ws3.freeze_panes = "A5"

# ---------- 04 KPI ----------
ws4 = wb.create_sheet("04_KPI_свод")
ws4["A1"] = "P&L-lite: свод по валовой прибыли (без opex / net)"
ws4["A1"].font = title_font
ws4["A2"] = "Верх: KPI-карточки. Низ: матрица месяц × группа номенклатуры. Модель PTM: только sales / COGS / gross."
ws4["A2"].font = subtitle_font
ws4.merge_cells("A2:F2")

tot_s = sum(r["sell"] for r in rows)
tot_c = sum(r["cost"] for r in rows)
tot_g = sum(r["gp"] for r in rows)
tot_r = tot_g / tot_s if tot_s else 0

kpis = [
    ("Продажи (выручка)", tot_s, money_fmt, "E2EFDA"),
    ("Себестоимость (COGS)", tot_c, money_fmt, "FCE4D6"),
    ("Валовая прибыль", tot_g, money_fmt, "DDEBF7"),
    ("Рентабельность %", tot_r, pct_fmt, "FFF2CC"),
]
for i, (label, val, fmt, color) in enumerate(kpis):
    col = 1 + i * 2
    ws4.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)
    ws4.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col + 1)
    c1 = ws4.cell(row=4, column=col, value=label)
    c1.font = kpi_label
    c1.fill = PatternFill("solid", fgColor=color)
    c1.alignment = Alignment(horizontal="center")
    c1.border = thin
    ws4.cell(row=4, column=col + 1).fill = PatternFill("solid", fgColor=color)
    ws4.cell(row=4, column=col + 1).border = thin
    c2 = ws4.cell(row=5, column=col, value=round(val, 4) if fmt == pct_fmt else round(val, 2))
    c2.font = kpi_value
    c2.number_format = fmt
    c2.fill = PatternFill("solid", fgColor=color)
    c2.alignment = Alignment(horizontal="center")
    c2.border = thin
    ws4.cell(row=5, column=col + 1).fill = PatternFill("solid", fgColor=color)
    ws4.cell(row=5, column=col + 1).border = thin
    ws4.row_dimensions[5].height = 28

ws4["A8"] = "Матрица: месяц × группа (валовая прибыль)"
ws4["A8"].font = Font(bold=True, size=12, color="1F4E79")

months = sorted({r["month"] for r in rows})
groups = sorted({r["group"] for r in rows})
mat = defaultdict(lambda: defaultdict(float))
for r in rows:
    mat[r["group"]][r["month"]] += r["gp"]

ws4.cell(row=9, column=1, value="Группа")
for j, m in enumerate(months, 2):
    ws4.cell(row=9, column=j, value=m)
ws4.cell(row=9, column=2 + len(months), value="Итого")
style_header(ws4, 9, 1 + len(months) + 1)

row_i = 10
for g in groups:
    ws4.cell(row=row_i, column=1, value=g)
    s = 0
    for j, m in enumerate(months, 2):
        v = round(mat[g][m], 2)
        s += v
        cell = ws4.cell(row=row_i, column=j, value=v)
        cell.number_format = money_fmt
    cell = ws4.cell(row=row_i, column=2 + len(months), value=round(s, 2))
    cell.number_format = money_fmt
    style_data_row(ws4, row_i, 1 + len(months) + 1)
    row_i += 1

ws4.cell(row=row_i, column=1, value="Итого").font = total_font
for j, m in enumerate(months, 2):
    v = round(sum(mat[g][m] for g in groups), 2)
    cell = ws4.cell(row=row_i, column=j, value=v)
    cell.number_format = money_fmt
cell = ws4.cell(row=row_i, column=2 + len(months), value=round(tot_g, 2))
cell.number_format = money_fmt
style_data_row(ws4, row_i, 1 + len(months) + 1, is_total=True)

last_col = get_column_letter(1 + len(months))
ws4.conditional_formatting.add(
    f"B10:{last_col}{row_i - 1}",
    ColorScaleRule(
        start_type="min",
        start_color="F8696B",
        mid_type="percentile",
        mid_value=50,
        mid_color="FFEB84",
        end_type="max",
        end_color="63BE7B",
    ),
)

row_i += 3
ws4.cell(row=row_i, column=1, value="Рентабельность % по группам (весь период)").font = Font(
    bold=True, size=12, color="1F4E79"
)
row_i += 1
for j, h in enumerate(["Группа", "Продажи", "Себестоимость", "Валовая прибыль", "Рентабельность %"], 1):
    ws4.cell(row=row_i, column=j, value=h)
style_header(ws4, row_i, 5)
row_i += 1

agg_g = defaultdict(lambda: {"cost": 0, "sell": 0, "gp": 0})
for r in rows:
    a = agg_g[r["group"]]
    a["cost"] += r["cost"]
    a["sell"] += r["sell"]
    a["gp"] += r["gp"]

start_chart = row_i
for g, a in sorted(agg_g.items(), key=lambda x: -x[1]["gp"]):
    rent = a["gp"] / a["sell"] if a["sell"] else 0
    vals = [g, round(a["sell"], 2), round(a["cost"], 2), round(a["gp"], 2), rent]
    for j, v in enumerate(vals, 1):
        cell = ws4.cell(row=row_i, column=j, value=v)
        if j in (2, 3, 4):
            cell.number_format = money_fmt
        if j == 5:
            cell.number_format = pct_fmt
    style_data_row(ws4, row_i, 5)
    row_i += 1

pie = PieChart()
pie.title = "Доля валовой прибыли по группам"
labels = Reference(ws4, min_col=1, min_row=start_chart, max_row=row_i - 1)
data = Reference(ws4, min_col=4, min_row=start_chart - 1, max_row=row_i - 1)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
pie.dataLabels = DataLabelList()
pie.dataLabels.showPercent = True
pie.dataLabels.showVal = False
pie.dataLabels.showCatName = False
pie.width = 14
pie.height = 10
ws4.add_chart(pie, "G14")

for col in range(1, 12):
    ws4.column_dimensions[get_column_letter(col)].width = 14
ws4.column_dimensions["A"].width = 18

# ---------- 05 raw ----------
ws5 = wb.create_sheet("05_Сырые_данные")
ws5["A1"] = "Сырые строки (для своей сводной таблицы Excel)"
ws5["A1"].font = title_font
ws5["A2"] = "Выделите таблицу → Вставка → Сводная таблица. Можно собрать любой срез."
ws5["A2"].font = subtitle_font

raw_h = [
    "Дата",
    "Месяц",
    "Номенклатура",
    "Группа",
    "Касса",
    "Сотрудник",
    "Количество",
    "Себестоимость",
    "Продажная сумма",
    "Валовая прибыль",
    "Рентабельность %",
]
for j, h in enumerate(raw_h, 1):
    ws5.cell(row=4, column=j, value=h)
style_header(ws5, 4, 11)

for i, r in enumerate(sorted(rows, key=lambda x: (x["date"], x["product"])), start=5):
    vals = [
        r["date"],
        r["month"],
        r["product"],
        r["group"],
        r["cashier"],
        r["staff"],
        r["qty"],
        r["cost"],
        r["sell"],
        r["gp"],
        r["r"],
    ]
    for j, v in enumerate(vals, 1):
        cell = ws5.cell(row=i, column=j, value=v)
        if j == 1:
            cell.number_format = "DD.MM.YYYY"
        if j in (8, 9, 10):
            cell.number_format = money_fmt
        if j == 7:
            cell.number_format = num_fmt
        if j == 11:
            cell.number_format = pct_fmt
    style_data_row(ws5, i, 11)

ws5.auto_filter.ref = f"A4:K{4 + len(rows)}"
ws5.freeze_panes = "A5"
auto_width(ws5)

wb.save(out)
print("OK", out)
print("rows", len(rows))
print("totals sell/cost/gp", round(tot_s, 2), round(tot_c, 2), round(tot_g, 2), f"{tot_r * 100:.1f}%")
