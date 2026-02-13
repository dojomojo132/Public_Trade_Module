# -*- coding: utf-8 -*-
import pathlib
import openpyxl

base = pathlib.Path(r"D:\Git\Public_Trade_Module\Import data")

if not base.exists():
    print(f"Path not found: {base}")
    import sys; sys.exit(1)

files = sorted(base.glob("*.xlsx"))
print(f"Found {len(files)} files")

for f in files:
    print(f"\n=== {f.name} ===")
    wb = openpyxl.load_workbook(str(f), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(max_row=5, values_only=True))
    for i, row in enumerate(rows):
        tag = "HEADER" if i == 0 else f"row {i+1}"
        vals = [str(v) if v is not None else "" for v in row]
        print(f"  [{tag}] {vals}")
    print(f"  Total rows: {ws.max_row}")
    wb.close()
    
print("\nDone!")
